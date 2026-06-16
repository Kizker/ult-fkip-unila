import openpyxl
from pathlib import Path

def main():
    val_path = Path(r"C:\laragon\www\ult-fkip-unila\docs\skripsi\hasil uji ahli validator\Rekap_Uji_Validitas_Ahli.xlsx")
    pr_path = Path(r"C:\laragon\www\ult-fkip-unila\docs\skripsi\hasil uji kepraktisan\Rekap_Uji_Kepraktisan.xlsx")
    
    if val_path.exists():
        print(f"=== {val_path.name} ===")
        wb = openpyxl.load_workbook(val_path, data_only=True)
        print("Sheets:", wb.sheetnames)
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\nSheet {sname} dimension: {ws.dimensions}")
            # print first 5 rows
            for idx, r in enumerate(ws.iter_rows(values_only=True)):
                if idx < 6:
                    print(f"Row {idx}: {r}")
                    
    if pr_path.exists():
        print(f"\n=== {pr_path.name} ===")
        wb = openpyxl.load_workbook(pr_path, data_only=True)
        print("Sheets:", wb.sheetnames)
        ws = wb.active
        print(f"Active Sheet dimension: {ws.dimensions}")
        for idx, r in enumerate(ws.iter_rows(values_only=True)):
            if idx < 10:
                print(f"Row {idx}: {r[:15]}") # first 15 columns

if __name__ == "__main__":
    main()
