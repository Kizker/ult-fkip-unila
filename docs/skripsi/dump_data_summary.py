import openpyxl
from pathlib import Path

def main():
    val_path = Path(r"C:\laragon\www\ult-fkip-unila\docs\skripsi\hasil uji ahli validator\Rekap_Uji_Validitas_Ahli.xlsx")
    pr_path = Path(r"C:\laragon\www\ult-fkip-unila\docs\skripsi\hasil uji kepraktisan\Rekap_Uji_Kepraktisan.xlsx")
    
    out_path = Path(r"C:\laragon\www\ult-fkip-unila\docs\skripsi\xlsx_data_summary.txt")
    
    with open(out_path, "w", encoding="utf-8") as f:
        if val_path.exists():
            f.write(f"=== VALIDITAS AHLI DATA ({val_path.name}) ===\n")
            wb = openpyxl.load_workbook(val_path, data_only=True)
            for sname in wb.sheetnames:
                f.write(f"\n--- Sheet: {sname} ---\n")
                ws = wb[sname]
                for r_idx, r in enumerate(ws.iter_rows(values_only=True)):
                    # skip empty rows
                    if any(x is not None for x in r):
                        vals = [str(x) if x is not None else "" for x in r]
                        f.write(f"Row {r_idx}: {vals}\n")
                        
        if pr_path.exists():
            f.write(f"\n=== KEPRAKTISAN DATA ({pr_path.name}) ===\n")
            wb = openpyxl.load_workbook(pr_path, data_only=True)
            for sname in wb.sheetnames:
                f.write(f"\n--- Sheet: {sname} ---\n")
                ws = wb[sname]
                for r_idx, r in enumerate(ws.iter_rows(values_only=True)):
                    if any(x is not None for x in r):
                        vals = [str(x) if x is not None else "" for x in r]
                        f.write(f"Row {r_idx}: {vals}\n")
                        
    print(f"Data summary dumped to {out_path}")

if __name__ == "__main__":
    main()
