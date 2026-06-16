import openpyxl
import os

def read_excel(file_path):
    print(f"--- Reading {os.path.basename(file_path)} ---")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"\nSheet: {sheet_name}")
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            # Only print rows that have some text that might be comments
            if any(isinstance(cell, str) and len(cell) > 10 for cell in row):
                print(row)

read_excel(r"C:\laragon\www\ult-fkip-unila\docs\skripsi\hasil uji ahli validator\Rekap_Uji_Validitas_Ahli.xlsx")
read_excel(r"C:\laragon\www\ult-fkip-unila\docs\skripsi\hasil uji kepraktisan\Rekap_Uji_Kepraktisan.xlsx")
