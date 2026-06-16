import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def generate_excel():
    json_path = "adaptive_omr_results.json"
    if not os.path.exists(json_path):
        print(f"Error: {json_path} does not exist.")
        return
        
    with open(json_path, "r") as f:
        results = json.load(f)
        
    # Maps for names and roles
    role_map = {
        "admin": "Admin",
        "pbs": "PBS (Pend. Bahasa & Sastra)",
        "pip": "PIP (Pend. Ilmu Pengetahuan)",
        "pips": "PIPS (Pend. Ilmu Pengetahuan Sosial)",
        "pmipa": "PMIPA (Pend. Matematika & IPA)",
        "ult": "ULT (Unit Layanan Terpadu)"
    }
    
    conclusion_map = {
        "a": "Sangat praktis untuk digunakan",
        "b": "Praktis untuk digunakan",
        "c": "Cukup praktis untuk digunakan",
        "d": "Kurang praktis untuk digunakan",
        "e": "Tidak praktis untuk digunakan"
    }
    
    # Initialize Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Uji Kepraktisan"
    ws.views.sheetView[0].showGridLines = True
    
    # Design colors
    HEADER_FILL = PatternFill(start_color="0F4C5C", end_color="0F4C5C", fill_type="solid") # Dark Teal
    ZEBRA_FILL = PatternFill(start_color="F4F9FA", end_color="F4F9FA", fill_type="solid")  # Ice Blue/Teal
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    TOTALS_FILL = PatternFill(start_color="EAF2F4", end_color="EAF2F4", fill_type="solid") # Muted Teal Accent
    
    # Fonts
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="0F4C5C")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    data_bold = Font(name=font_family, size=10, bold=True, color="0F4C5C")
    totals_font = Font(name=font_family, size=11, bold=True, color="000000")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D0D9DC")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom_border = Border(
        top=Side(border_style="thin", color="0F4C5C"),
        bottom=Side(border_style="double", color="0F4C5C"),
        left=thin_border_side,
        right=thin_border_side
    )
    
    # Write Title Block (Rows 1 to 5)
    ws["B2"] = "REKAPITULASI HASIL UJI KEPRAKTISAN WEBSITE"
    ws["B2"].font = title_font
    ws["B3"] = "Unit Layanan Terpadu (ULT) FKIP Universitas Lampung"
    ws["B3"].font = title_font
    ws["B4"] = "Instrumen Uji Kepraktisan - Skala Likert 1-5 (5 = Sangat Setuju, 1 = Sangat Tidak Setuju)"
    ws["B4"].font = subtitle_font
    
    # Header Row (Row 7)
    headers = [
        "No", "Nama Responden", "Peran / Program Studi",
        "A1", "A2", "A3", "B1", "B2", "B3", "C1", "C2", "D1", "D2", "E1", "E2",
        "Total Skor", "Persentase (%)", "Kategori Kepraktisan", "Kesimpulan Responden"
    ]
    
    header_row = 7
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    start_row = 8
    
    # Populate Data
    for idx, (pdf_path, res) in enumerate(sorted(results.items()), 1):
        r = start_row + idx - 1
        
        # Parse filename to get role and name
        filename = os.path.basename(pdf_path).replace(".pdf", "")
        parts = filename.split(" ")
        role_raw = parts[2].lower()
        role = role_map.get(role_raw, role_raw.upper())
        
        name_raw = " ".join([p for p in parts[3:] if p.strip()])
        name = name_raw.strip().title()
        
        scores = res["scores"]
        conclusion_code = res["conclusion"]
        conclusion = conclusion_map.get(conclusion_code, "Tidak valid")
        
        # Fill standard values
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=role)
        ws.cell(row=r, column=4, value=scores.get("A1", 0))
        ws.cell(row=r, column=5, value=scores.get("A2", 0))
        ws.cell(row=r, column=6, value=scores.get("A3", 0))
        ws.cell(row=r, column=7, value=scores.get("B1", 0))
        ws.cell(row=r, column=8, value=scores.get("B2", 0))
        ws.cell(row=r, column=9, value=scores.get("B3", 0))
        ws.cell(row=r, column=10, value=scores.get("C1", 0))
        ws.cell(row=r, column=11, value=scores.get("C2", 0))
        ws.cell(row=r, column=12, value=scores.get("D1", 0))
        ws.cell(row=r, column=13, value=scores.get("D2", 0))
        ws.cell(row=r, column=14, value=scores.get("E1", 0))
        ws.cell(row=r, column=15, value=scores.get("E2", 0))
        ws.cell(row=r, column=19, value=conclusion)
        
        # Formatting & Formulas
        row_fill = ZEBRA_FILL if idx % 2 == 0 else WHITE_FILL
        
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            if c in [1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        # Total Skor Formula: SUM of columns D to O (col 4 to 15)
        total_cell = ws.cell(row=r, column=16, value=f"=SUM(D{r}:O{r})")
        total_cell.font = data_bold
        total_cell.fill = row_fill
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Percentage Formula: Total Skor / 60 * 100
        pct_cell = ws.cell(row=r, column=17, value=f"=P{r}/60*100")
        pct_cell.font = data_bold
        pct_cell.fill = row_fill
        pct_cell.border = thin_border
        pct_cell.number_format = "0.00"
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Kategori Kepraktisan Formula (Dynamically mapped from Kesimpulan Responden in Column S):
        # "Sangat praktis untuk digunakan" -> "Sangat Praktis"
        # "Praktis untuk digunakan" -> "Praktis"
        # "Cukup praktis untuk digunakan" -> "Cukup Praktis"
        # "Kurang praktis untuk digunakan" -> "Kurang Praktis"
        # "Tidak praktis untuk digunakan" -> "Tidak Praktis"
        cat_formula = f'=IF(LEFT(S{r}, 6)="Sangat", "Sangat Praktis", IF(LEFT(S{r}, 7)="Praktis", "Praktis", IF(LEFT(S{r}, 5)="Cukup", "Cukup Praktis", IF(LEFT(S{r}, 6)="Kurang", "Kurang Praktis", "Tidak Praktis"))))'
        cat_cell = ws.cell(row=r, column=18, value=cat_formula)
        cat_cell.font = data_bold
        cat_cell.fill = row_fill
        cat_cell.border = thin_border
        cat_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Kesimpulan Responden Styling
        con_cell = ws.cell(row=r, column=19)
        con_cell.font = data_font
        con_cell.fill = row_fill
        con_cell.border = thin_border
        con_cell.alignment = Alignment(horizontal="left", vertical="center")
        
    end_row = start_row + len(results) - 1
    totals_row = end_row + 1
    
    # --- ADD TOTALS / AVERAGE ROW ---
    ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=3)
    totals_label = ws.cell(row=totals_row, column=1, value="Rata-rata Skor / Persentase")
    totals_label.font = totals_font
    totals_label.fill = TOTALS_FILL
    totals_label.alignment = Alignment(horizontal="right", vertical="center")
    totals_label.border = double_bottom_border
    
    # Border for the merged cells
    for c in range(1, 4):
        ws.cell(row=totals_row, column=c).border = double_bottom_border
        ws.cell(row=totals_row, column=c).fill = TOTALS_FILL
        
    # Average for individual indicators A1-E2
    for c in range(4, 16):
        col_let = get_column_letter(c)
        avg_cell = ws.cell(row=totals_row, column=c, value=f"=AVERAGE({col_let}{start_row}:{col_let}{end_row})")
        avg_cell.font = totals_font
        avg_cell.fill = TOTALS_FILL
        avg_cell.border = double_bottom_border
        avg_cell.number_format = "0.00"
        avg_cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Average Total Skor
    avg_total = ws.cell(row=totals_row, column=16, value=f"=AVERAGE(P{start_row}:P{end_row})")
    avg_total.font = totals_font
    avg_total.fill = PatternFill(start_color="D6E5E8", end_color="D6E5E8", fill_type="solid") # Darker totals accent
    avg_total.border = double_bottom_border
    avg_total.number_format = "0.00"
    avg_total.alignment = Alignment(horizontal="center", vertical="center")
    
    # Average Percentage
    avg_pct = ws.cell(row=totals_row, column=17, value=f"=AVERAGE(Q{start_row}:Q{end_row})")
    avg_pct.font = totals_font
    avg_pct.fill = PatternFill(start_color="C2D9DE", end_color="C2D9DE", fill_type="solid") # Even darker accent
    avg_pct.border = double_bottom_border
    avg_pct.number_format = "0.00"
    avg_pct.alignment = Alignment(horizontal="center", vertical="center")
    
    # Overall Category Formula
    overall_cat = ws.cell(row=totals_row, column=18, value=f'=IF(Q{totals_row}>=81, "Sangat Praktis", IF(Q{totals_row}>=61, "Praktis", IF(Q{totals_row}>=41, "Cukup Praktis", IF(Q{totals_row}>=21, "Kurang Praktis", "Tidak Praktis"))))')
    overall_cat.font = totals_font
    overall_cat.fill = PatternFill(start_color="C2D9DE", end_color="C2D9DE", fill_type="solid")
    overall_cat.border = double_bottom_border
    overall_cat.alignment = Alignment(horizontal="center", vertical="center")
    
    # Empty for conclusion in totals row
    empty_con = ws.cell(row=totals_row, column=19, value="-")
    empty_con.font = totals_font
    empty_con.fill = TOTALS_FILL
    empty_con.border = double_bottom_border
    empty_con.alignment = Alignment(horizontal="center", vertical="center")
    
    # --- ADD DISTRIBUTIONS SUMMARY BLOCK ---
    dist_start_row = totals_row + 3
    
    ws.cell(row=dist_start_row, column=2, value="RINGKASAN DISTRIBUSI KATEGORI").font = Font(name=font_family, size=11, bold=True, color="0F4C5C")
    
    categories_dist = [
        ("Sangat Praktis (81% - 100%)", f'=COUNTIF(R{start_row}:R{end_row}, "Sangat Praktis")'),
        ("Praktis (61% - 80%)", f'=COUNTIF(R{start_row}:R{end_row}, "Praktis")'),
        ("Cukup Praktis (41% - 60%)", f'=COUNTIF(R{start_row}:R{end_row}, "Cukup Praktis")'),
        ("Kurang Praktis (21% - 40%)", f'=COUNTIF(R{start_row}:R{end_row}, "Kurang Praktis")'),
        ("Tidak Praktis (0% - 20%)", f'=COUNTIF(R{start_row}:R{end_row}, "Tidak Praktis")'),
    ]
    
    for idx, (label, formula) in enumerate(categories_dist):
        curr_row = dist_start_row + 1 + idx
        
        lbl_cell = ws.cell(row=curr_row, column=2, value=label)
        lbl_cell.font = data_font
        lbl_cell.border = thin_border
        
        val_cell = ws.cell(row=curr_row, column=3, value=formula)
        val_cell.font = data_bold
        val_cell.border = thin_border
        val_cell.alignment = Alignment(horizontal="center")
        
    # Total Responden in Distribution Block
    total_resp_row = dist_start_row + 1 + len(categories_dist)
    lbl_cell = ws.cell(row=total_resp_row, column=2, value="Total Responden")
    lbl_cell.font = Font(name=font_family, size=10, bold=True)
    lbl_cell.border = thin_border
    lbl_cell.fill = TOTALS_FILL
    
    val_cell = ws.cell(row=total_resp_row, column=3, value=f"=SUM(C{dist_start_row+1}:C{total_resp_row-1})")
    val_cell.font = Font(name=font_family, size=10, bold=True)
    val_cell.border = thin_border
    val_cell.fill = TOTALS_FILL
    val_cell.alignment = Alignment(horizontal="center")
    
    # Auto-fit Column Widths (with padding)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row > 6:
                val_str = str(cell.value or "")
                if col_letter == "S": # Conclusion column
                    max_len = max(max_len, 25) # Standard size
                elif col_letter == "B": # Name column
                    max_len = max(max_len, len(val_str))
                elif col_letter == "C": # Role column
                    max_len = max(max_len, len(val_str))
                else:
                    max_len = max(max_len, len(val_str))
                    
        ws.column_dimensions[col_letter].width = max(max_len + 4, 8)
        
    # Specific adjustments
    ws.column_dimensions["S"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    
    # Set Row Heights
    ws.row_dimensions[7].height = 28 # Header
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 20 # Data
    ws.row_dimensions[totals_row].height = 24 # Totals
    
    output_excel = "Rekap_Uji_Kepraktisan.xlsx"
    wb.save(output_excel)
    print(f"Successfully generated pure openpyxl styled Excel: {output_excel}")

if __name__ == "__main__":
    generate_excel()
