"""
Script untuk menambahkan kolom 'Komentar/Saran Perbaikan' ke Rekap_Uji_Kepraktisan.xlsx
Data komentar berasal dari Rekap_Komentar_Saran.pdf (sudah diekstrak sebelumnya).
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import shutil

def add_komentar_kolom():
    excel_path = "Rekap_Uji_Kepraktisan.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} tidak ditemukan.")
        return
    
    # Backup file asli
    backup_path = "Rekap_Uji_Kepraktisan_BACKUP.xlsx"
    shutil.copy2(excel_path, backup_path)
    print(f"Backup dibuat: {backup_path}")
    
    # Data komentar/saran per responden (diekstrak dari Rekap_Komentar_Saran.pdf)
    # Key = nama responden (lowercase, stripped), Value = komentar/saran
    komentar_data = {
        "anisa": "Mungkin bisa ditambahkan fitur untuk melihat jadwal mahasiswa seminar, seperti airtable.com",
        "lisa": "Website ini sudah cukup bagus untuk membantu proses administrasi di lingkungan FKIP. Saran untuk penomoran agar dapat langsung me-link ke penomoran fakultas.",
        "riswan": "Mudah dan segera berjalan & bisa dimanfaatkan untuk kelancaran proses administrasi, juga untuk mahasiswa.",
        "khaerul": "(Tidak memberikan komentar/saran)",
        "martin": "(Tidak memberikan komentar/saran)",
        "nurani": "(Tidak memberikan komentar/saran)",
        "aulia": "(Tidak memberikan komentar/saran)",
        "nazwa": "(Tidak memberikan komentar/saran)",
        "salsa": "(Tidak memberikan komentar/saran)",
        "andhini": "(Tidak memberikan komentar/saran)",
        "arya": "(Tidak memberikan komentar/saran)",
        "mita": "(Tidak memberikan komentar/saran)",
        "nabila": "Display website nya simple tapi eye catching, informasi yang disediakan juga lengkap & jelas. Perkiraan saya jika memang nanti bisa digunakan secara real akan sangat membantu proses administrasi mahasiswa.",
        "nur": "Susunan teksnya jangan terlalu banyak sehingga terkesan menumpuk.",
        "rizky": "(Tidak memberikan komentar/saran)",
        "agus": "(Tidak memberikan komentar/saran)",
        "amrul": "1. Grafik Transaksi Data Bulanan/Mingguan (Pelaporan tabel). 2. Tambahkan Menu Layanan Kepegawaian.",
        "tri": "(Tidak memberikan komentar/saran)",
    }
    
    # Buka workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Design styles (konsisten dengan skema Teal/Sage yang sudah ada)
    HEADER_FILL = PatternFill(start_color="0F4C5C", end_color="0F4C5C", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F4F9FA", end_color="F4F9FA", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    TOTALS_FILL = PatternFill(start_color="EAF2F4", end_color="EAF2F4", fill_type="solid")
    
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    data_font_italic = Font(name=font_family, size=10, color="888888", italic=True)
    totals_font = Font(name=font_family, size=11, bold=True, color="000000")
    
    thin_border_side = Side(border_style="thin", color="D0D9DC")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(
        top=Side(border_style="thin", color="0F4C5C"),
        bottom=Side(border_style="double", color="0F4C5C"),
        left=thin_border_side,
        right=thin_border_side
    )
    
    # Kolom baru = kolom T (20) -> "Komentar / Saran Perbaikan"
    new_col = 20  # Kolom T
    header_row = 7
    start_row = 8
    num_respondents = 18
    end_row = start_row + num_respondents - 1  # 25
    totals_row = end_row + 1  # 26
    
    # Tulis header baru
    header_cell = ws.cell(row=header_row, column=new_col, value="Komentar / Saran Perbaikan")
    header_cell.font = header_font
    header_cell.fill = HEADER_FILL
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_cell.border = thin_border
    
    # Isi data komentar per responden
    matched_count = 0
    for r in range(start_row, end_row + 1):
        # Baca nama responden dari kolom B (kolom 2)
        nama_cell = ws.cell(row=r, column=2).value
        if nama_cell is None:
            continue
            
        nama_key = str(nama_cell).strip().lower()
        
        # Cari komentar yang cocok (exact match dulu, baru substring)
        komentar = None
        # 1. Exact match
        if nama_key in komentar_data:
            komentar = komentar_data[nama_key]
            matched_count += 1
        else:
            # 2. Substring fallback
            for key, val in komentar_data.items():
                if key in nama_key or nama_key in key:
                    komentar = val
                    matched_count += 1
                    break
        
        if komentar is None:
            komentar = "(Tidak memberikan komentar/saran)"
        
        # Tulis ke kolom T
        cell = ws.cell(row=r, column=new_col, value=komentar)
        
        # Styling
        idx = r - start_row + 1
        row_fill = ZEBRA_FILL if idx % 2 == 0 else WHITE_FILL
        
        # Gunakan font italic abu-abu untuk yang tidak ada komentar
        is_empty = "(Tidak memberikan" in komentar
        cell.font = data_font_italic if is_empty else data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Isi totals row (baris rata-rata) untuk kolom komentar
    totals_cell = ws.cell(row=totals_row, column=new_col)
    totals_cell.value = f"Memberikan komentar: 6 dari 18 responden"
    totals_cell.font = totals_font
    totals_cell.fill = TOTALS_FILL
    totals_cell.border = double_bottom_border
    totals_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Atur lebar kolom T agar muat untuk teks komentar
    ws.column_dimensions[get_column_letter(new_col)].width = 55
    
    # Atur tinggi baris yang memiliki komentar panjang agar wrap_text terlihat baik
    for r in range(start_row, end_row + 1):
        komentar_val = ws.cell(row=r, column=new_col).value or ""
        if len(komentar_val) > 60 and "(Tidak memberikan" not in komentar_val:
            # Estimasi tinggi baris berdasarkan panjang teks
            est_lines = max(2, len(komentar_val) // 50 + 1)
            current_height = ws.row_dimensions[r].height or 20
            new_height = max(current_height, est_lines * 15)
            ws.row_dimensions[r].height = new_height
    
    # Simpan
    wb.save(excel_path)
    print(f"\n{'='*60}")
    print(f"SUKSES: Kolom 'Komentar / Saran Perbaikan' berhasil ditambahkan!")
    print(f"{'='*60}")
    print(f"File: {excel_path}")
    print(f"Kolom baru: T (kolom ke-20)")
    print(f"Responden yang di-match: {matched_count} dari {num_respondents}")
    print(f"Responden memberikan komentar: 6 orang")
    print(f"Responden tanpa komentar: 12 orang")
    print(f"\nResponden yang memberikan komentar/saran:")
    for name in ["Anisa", "Lisa", "Riswan", "Nabila", "Nur", "Amrul"]:
        kmt = komentar_data[name.lower()]
        print(f"  - {name}: \"{kmt[:60]}{'...' if len(kmt) > 60 else ''}\"")
    print(f"\nBackup file asli: {backup_path}")

if __name__ == "__main__":
    add_komentar_to_excel = add_komentar_kolom
    add_komentar_to_excel()
