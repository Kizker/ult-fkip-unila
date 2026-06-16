import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

def apply_formatting_and_formulas(filename):
    wb = load_workbook(filename)
    
    # Sleek Color Palette
    header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid") # Dark Slate Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF") # White text
    
    sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue Grey
    sub_header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), 
                         top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))

    number_format_1_decimal = "0.0"

    # Soft HSL style colors for conditional formatting
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    green_font = Font(name="Calibri", color="375623", bold=True)
    
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # soft blue
    blue_font = Font(name="Calibri", color="1F4E78", bold=True)
    
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # soft yellow
    yellow_font = Font(name="Calibri", color="7F6000", bold=True)
    
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft red
    red_font = Font(name="Calibri", color="C65911", bold=True)

    def format_sheet(ws, num_items):
        max_row = num_items + 1  
        
        # 1. Populate Formulas for items
        for row in range(2, max_row + 1):
            ws[f'F{row}'] = f'=SUM(C{row}:E{row})'
            ws[f'G{row}'] = f'=ROUND(AVERAGE(C{row}:E{row}), 0)'
            ws[f'H{row}'] = 15
            ws[f'I{row}'] = f'=(F{row}/H{row})*100'
            ws[f'I{row}'].number_format = number_format_1_decimal
            ws[f'J{row}'] = f'=IF(I{row}>=81,"Sangat Valid",IF(I{row}>=61,"Valid",IF(I{row}>=41,"Cukup Valid",IF(I{row}>=21,"Kurang Valid","Tidak Valid"))))'
        
        # 2. Format Table (Header & Cells)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                else:
                    cell.font = Font(name="Calibri", size=11)
                    if cell.column in [1, 2]:  # Aspek and Indikator
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                        
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 45
        for col_idx in range(3, 11):
            column = get_column_letter(col_idx)
            ws.column_dimensions[column].width = 15

        ws.freeze_panes = 'C2'
        
        # Conditional Formatting
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='equal', formula=['"Sangat Valid"'], stopIfTrue=True, fill=green_fill, font=green_font))
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='equal', formula=['"Valid"'], stopIfTrue=True, fill=blue_fill, font=blue_font))
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='equal', formula=['"Cukup Valid"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='equal', formula=['"Kurang Valid"'], stopIfTrue=True, fill=red_fill, font=red_font))
        ws.conditional_formatting.add(f'J2:J{max_row}', CellIsRule(operator='equal', formula=['"Tidak Valid"'], stopIfTrue=True, fill=red_fill, font=red_font))

    format_sheet(wb['Validitas Materi'], 22)
    format_sheet(wb['Validitas Media'], 14)
    format_sheet(wb['Validitas Sistem'], 22)

    # ==========================
    # FORMAT REKAP KESELURUHANN
    # ==========================
    ws_rekap = wb['Rekap Keseluruhan']
    
    # Materi (Baris 2)
    ws_rekap['B2'] = "='Validitas Materi'!F24"
    ws_rekap['C2'] = "='Validitas Materi'!G24"  
    ws_rekap['D2'] = "='Validitas Materi'!H24"  
    ws_rekap['E2'] = "=(B2/D2)*100"
    ws_rekap['F2'] = '=IF(E2>=81,"Sangat Valid",IF(E2>=61,"Valid",IF(E2>=41,"Cukup Valid",IF(E2>=21,"Kurang Valid","Tidak Valid"))))'
    
    # Media (Baris 3)
    ws_rekap['B3'] = "='Validitas Media'!F16"
    ws_rekap['C3'] = "='Validitas Media'!G16"
    ws_rekap['D3'] = "='Validitas Media'!H16"
    ws_rekap['E3'] = "=(B3/D3)*100"
    ws_rekap['F3'] = '=IF(E3>=81,"Sangat Valid",IF(E3>=61,"Valid",IF(E3>=41,"Cukup Valid",IF(E3>=21,"Kurang Valid","Tidak Valid"))))'
    
    # Sistem (Baris 4)
    ws_rekap['B4'] = "='Validitas Sistem'!F24"
    ws_rekap['C4'] = "='Validitas Sistem'!G24"
    ws_rekap['D4'] = "='Validitas Sistem'!H24"
    ws_rekap['E4'] = "=(B4/D4)*100"
    ws_rekap['F4'] = '=IF(E4>=81,"Sangat Valid",IF(E4>=61,"Valid",IF(E4>=41,"Cukup Valid",IF(E4>=21,"Kurang Valid","Tidak Valid"))))'

    # Rata-rata Keseluruhan (Baris 5)
    ws_rekap['B5'] = "=SUM(B2:B4)"
    ws_rekap['C5'] = "=ROUND(AVERAGE(C2:C4),0)"
    ws_rekap['D5'] = "=SUM(D2:D4)"
    ws_rekap['E5'] = "=(B5/D5)*100"
    ws_rekap['F5'] = '=IF(E5>=81,"Sangat Valid",IF(E5>=61,"Valid",IF(E5>=41,"Cukup Valid",IF(E5>=21,"Kurang Valid","Tidak Valid"))))'

    for row_idx in range(2, 6):
        ws_rekap[f'E{row_idx}'].number_format = number_format_1_decimal

    # Format table 1 (Rekap Atas)
    for row in ws_rekap.iter_rows(min_row=1, max_row=5, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            else:
                cell.font = Font(name="Calibri", size=11)
                if cell.column == 1:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                if cell.row == 5:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    cell.fill = sub_header_fill

    # Format tabel 2 (Keputusan Bawah) - RANGING dari Row 9 s/d 21 secara utuh
    for row in ws_rekap.iter_rows(min_row=9, max_row=21, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11)
            if cell.row == 9:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            elif cell.column == 4:
                cell.alignment = left_align # Komentar/Saran left aligned
            else:
                cell.alignment = center_align

    # Menambahkan Mayoritas (Modus) Kelayakan di Rekap Keseluruhan secara benar (indeks 10, 14, 18)
    for start_r in [10, 14, 18]:
        target_row = start_r + 3
        ws_rekap.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=2)
        ws_rekap.cell(row=target_row, column=1, value="HASIL MAYORITAS KELAYAKAN").font = Font(name="Calibri", size=11, bold=True)
        ws_rekap.cell(row=target_row, column=1).alignment = left_align
        
        # Kolom C (keputusan)
        r1, r2, r3 = start_r, start_r+1, start_r+2
        ws_rekap.cell(row=target_row, column=3, value=f'=IF(C{r1}=C{r2},C{r1},IF(C{r2}=C{r3},C{r2},IF(C{r1}=C{r3},C{r1},"Draw")))')
        ws_rekap.cell(row=target_row, column=3).font = Font(name="Calibri", size=11, bold=True)
        ws_rekap.cell(row=target_row, column=3).alignment = center_align
        ws_rekap.cell(row=target_row, column=3).fill = green_fill
        
        ws_rekap.merge_cells(start_row=target_row, start_column=3, end_row=target_row, end_column=4)
        
        # Set border dan formatting untuk baris yang dimerge
        for c_idx in range(1, 5):
            cell = ws_rekap.cell(row=target_row, column=c_idx)
            cell.border = thin_border
            if c_idx >= 3:
                cell.fill = green_fill
                cell.font = green_font

    # Set Column widths rekap keseluruhan
    ws_rekap.column_dimensions['A'].width = 28  # Nama Validator
    ws_rekap.column_dimensions['B'].width = 18  # Aspek
    ws_rekap.column_dimensions['C'].width = 30  # Keputusan Kelayakan
    ws_rekap.column_dimensions['D'].width = 55  # Komentar / Saran
    ws_rekap.column_dimensions['E'].width = 15  # Persentase
    ws_rekap.column_dimensions['F'].width = 18  # Kriteria Validitas

    # CF untuk Kriteria Validitas Rekap
    ws_rekap.conditional_formatting.add('F2:F5', CellIsRule(operator='equal', formula=['"Sangat Valid"'], stopIfTrue=True, fill=green_fill, font=green_font))
    ws_rekap.conditional_formatting.add('F2:F5', CellIsRule(operator='equal', formula=['"Valid"'], stopIfTrue=True, fill=blue_fill, font=blue_font))
    ws_rekap.conditional_formatting.add('F2:F5', CellIsRule(operator='equal', formula=['"Cukup Valid"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
    ws_rekap.conditional_formatting.add('F2:F5', CellIsRule(operator='equal', formula=['"Kurang Valid"'], stopIfTrue=True, fill=red_fill, font=red_font))
    ws_rekap.conditional_formatting.add('F2:F5', CellIsRule(operator='equal', formula=['"Tidak Valid"'], stopIfTrue=True, fill=red_fill, font=red_font))

    # ==============================================================
    # TOTAL KESELURUHAN & KEPUTUSAN KELAYAKAN di Tiap Sheet Detail
    # ==============================================================
    def add_total_and_keputusan(ws, num_items, keputusan_data):
        row = num_items + 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws[f'A{row}'] = 'TOTAL KESELURUHAN'
        ws[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
        ws[f'A{row}'].alignment = left_align
        
        ws[f'C{row}'] = f'=SUM(C2:C{num_items+1})'
        ws[f'D{row}'] = f'=SUM(D2:D{num_items+1})'
        ws[f'E{row}'] = f'=SUM(E2:E{num_items+1})'
        ws[f'F{row}'] = f'=SUM(F2:F{num_items+1})'
        ws[f'G{row}'] = f'=ROUND(AVERAGE(G2:G{num_items+1}),0)'
        ws[f'H{row}'] = f'=SUM(H2:H{num_items+1})'
        ws[f'I{row}'] = f'=(F{row}/H{row})*100'
        ws[f'I{row}'].number_format = number_format_1_decimal
        ws[f'J{row}'] = f'=IF(I{row}>=81,"Sangat Valid",IF(I{row}>=61,"Valid",IF(I{row}>=41,"Cukup Valid",IF(I{row}>=21,"Kurang Valid","Tidak Valid"))))'
        
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.fill = sub_header_fill
            if col_idx > 2:
                cell.alignment = center_align

        # Sub-tabel Keputusan Kelayakan
        start_row_keputusan = row + 3
        
        # 1. Title baris
        ws.merge_cells(start_row=start_row_keputusan, start_column=1, end_row=start_row_keputusan, end_column=6)
        ws.cell(row=start_row_keputusan, column=1, value="HASIL KEPUTUSAN KELAYAKAN VALIDATOR").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=start_row_keputusan, column=1).alignment = left_align
        for c_idx in range(1, 7):
            ws.cell(row=start_row_keputusan, column=c_idx).border = thin_border
            ws.cell(row=start_row_keputusan, column=c_idx).fill = sub_header_fill
        
        # 2. Header kolom
        start_row_keputusan += 1
        ws.cell(row=start_row_keputusan, column=1, value="Nama Validator")
        ws.cell(row=start_row_keputusan, column=2, value="Keputusan Kelayakan")
        ws.merge_cells(start_row=start_row_keputusan, start_column=3, end_row=start_row_keputusan, end_column=6)
        ws.cell(row=start_row_keputusan, column=3, value="Komentar / Saran")
        
        for c_idx in range(1, 7):
            cell = ws.cell(row=start_row_keputusan, column=c_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # 3. Data baris (3 Validator)
        for i, data in enumerate(keputusan_data):
            r = start_row_keputusan + 1 + i
            ws.cell(row=r, column=1, value=data['nama']).alignment = center_align
            ws.cell(row=r, column=2, value=data['keputusan']).alignment = center_align
            
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            ws.cell(row=r, column=3, value=data['saran']).alignment = left_align
            
            for c_idx in range(1, 7):
                cell = ws.cell(row=r, column=c_idx)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=11)
                if c_idx >= 3:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                    
        # 4. KEPUTUSAN MAYORITAS (akhir)
        mayoritas_row = start_row_keputusan + 4
        ws.merge_cells(start_row=mayoritas_row, start_column=1, end_row=mayoritas_row, end_column=2)
        ws.cell(row=mayoritas_row, column=1, value="KEPUTUSAN KELAYAKAN AKHIR (MAYORITAS)").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=mayoritas_row, column=1).alignment = left_align
        
        r1, r2, r3 = start_row_keputusan + 1, start_row_keputusan + 2, start_row_keputusan + 3
        ws.cell(row=mayoritas_row, column=3, value=f'=IF(B{r1}=B{r2},B{r1},IF(B{r2}=B{r3},B{r2},IF(B{r1}=B{r3},B{r1},"Draw")))')
        ws.cell(row=mayoritas_row, column=3).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=mayoritas_row, column=3).alignment = center_align
        ws.cell(row=mayoritas_row, column=3).fill = green_fill
        
        ws.merge_cells(start_row=mayoritas_row, start_column=3, end_row=mayoritas_row, end_column=6)
        
        for c_idx in range(1, 7):
            cell = ws.cell(row=mayoritas_row, column=c_idx)
            cell.border = thin_border
            if c_idx >= 3:
                cell.fill = green_fill
                cell.font = green_font

    keputusan_materi = [
        {'nama': 'Margaretha Karolina S', 'keputusan': 'Layak digunakan dengan revisi', 'saran': 'Tambah penjelasan, cek kalimat, pastikan screenshot jelas'},
        {'nama': 'Eko Indra Pangestu', 'keputusan': 'Layak digunakan tanpa revisi', 'saran': '-'},
        {'nama': 'Putut Aji Nalendro', 'keputusan': 'Layak digunakan tanpa revisi', 'saran': 'Beri pembatas header/content/footer & gambar'}
    ]
    keputusan_media = [
        {'nama': 'Rafiqa Rizalita', 'keputusan': 'Layak digunakan tanpa revisi', 'saran': 'Bagus, bisa digunakan untuk penelitian'},
        {'nama': 'Daniel Rinaldi', 'keputusan': 'Layak digunakan tanpa revisi', 'saran': 'Warna sebaiknya jangan gradasi'},
        {'nama': 'Dwi Wahyudi', 'keputusan': 'Layak digunakan dengan revisi', 'saran': 'Beri icon pada setiap tile layanan'}
    ]
    keputusan_sistem = [
        {'nama': 'Ghea Chandra S', 'keputusan': 'Layak digunakan tanpa revisi', 'saran': 'Perbaikan saran'},
        {'nama': 'Radinal Fadli', 'keputusan': 'Layak digunakan dengan revisi', 'saran': 'Login SSO, perhatikan CSP (unsafe-eval/inline), atur CSP default'},
        {'nama': 'Rahmad Wahyudi', 'keputusan': 'Layak digunakan tanpa revisi', 'saran': 'Tampilkan tombol kembali ke atas'}
    ]

    add_total_and_keputusan(wb['Validitas Materi'], 22, keputusan_materi)
    add_total_and_keputusan(wb['Validitas Media'], 14, keputusan_media)
    add_total_and_keputusan(wb['Validitas Sistem'], 22, keputusan_sistem)

    wb.save(filename)

def main():
    aspek_materi = ['Kesesuaian Tujuan dan Ruang Lingkup']*4 + ['Akurasi dan Kelengkapan Isi']*8 + ['Keruntutan dan Kejelasan Penyajian']*6 + ['Kesesuaian Materi Pendukung']*4
    ind_materi = [
        'Kesesuaian tujuan panduan', 'Kesesuaian sasaran pengguna', 'Kesesuaian ruang lingkup', 'Kesesuaian ruang lingkup (fokus)',
        'Akurasi langkah penggunaan', 'Akurasi fitur dan istilah', 'Akurasi istilah konsisten', 'Akurasi informasi layanan',
        'Kelengkapan tahapan', 'Kelengkapan kondisi khusus', 'Kelengkapan informasi tindakan', 'Kelengkapan materi inti',
        'Keruntutan penyajian', 'Keruntutan antarbagian', 'Transisi pembahasan', 'Kejelasan bahasa', 'Kejelasan instruksi', 'Keringkasan penjelasan',
        'Relevansi materi pendukung', 'Kesesuaian ilustrasi', 'Kejelasan fungsi ilustrasi', 'Kesesuaian caption atau penjelas'
    ]
    materi_bu_etha = [5,5,5,5, 5,4,5,5,5,4,5,5, 5,5,5,5,4,5, 5,4,5,5]
    materi_pak_eko = [5,5,5,4, 5,4,5,5,5,4,4,4, 5,5,4,5,5,5, 5,5,5,5]
    materi_pak_putut = [5,5,5,5, 5,5,4,5,4,4,5,5, 5,5,5,4,5,5, 4,5,5,5]
    
    materi_data = {
        'Aspek Penilaian': aspek_materi,
        'Indikator Penilaian': ind_materi,
        'Margaretha Karolina S': materi_bu_etha,
        'Eko Indra Pangestu': materi_pak_eko,
        'Putut Aji Nalendro': materi_pak_putut,
        'Total Skor per Item': ["" for _ in range(22)],
        'Rata-rata Skor': ["" for _ in range(22)],
        'Skor Maks': ["" for _ in range(22)],
        'Persentase': ["" for _ in range(22)],
        'Kriteria Validitas': ["" for _ in range(22)]
    }
    df_materi = pd.DataFrame(materi_data)

    aspek_media = ['Tampilan Visual Website']*6 + ['UI/UX Website']*8
    ind_media = [
        'Kualitas tampilan website (rapi/menarik)', 'Kualitas tampilan website (warna)', 'Keterbacaan teks website (ukuran/huruf)', 
        'Keterbacaan teks website (kontras)', 'Konsistensi visual website (tata letak/ikon)', 'Konsistensi visual website (pengenalan)',
        'Kemudahan Navigasi', 'Kejelasan Interaksi & Media', 'Kualitas Audio-Visual Sistem', 'Umpan Balik Antarmuka', 
        'Efisiensi Alur Pelayanan', 'Kemudahan Penggunaan', 'Kesesuaian dengan Target', 'Konsistensi Pengalaman'
    ]
    media_bu_rafiqa = [5,4,4,5,5,4, 5,5,5,4,4,5,5,5]
    media_pak_daniel = [5,5,5,4,5,4, 5,4,5,5,4,5,4,5]
    media_pak_dwi = [5,5,5,5,5,5, 5,4,5,4,5,4,4,5]
    
    media_data = {
        'Aspek Penilaian': aspek_media,
        'Indikator Penilaian': ind_media,
        'Rafiqa Rizalita': media_bu_rafiqa,
        'Daniel Rinaldi': media_pak_daniel,
        'Dwi Wahyudi': media_pak_dwi,
        'Total Skor per Item': ["" for _ in range(14)],
        'Rata-rata Skor': ["" for _ in range(14)],
        'Skor Maks': ["" for _ in range(14)],
        'Persentase': ["" for _ in range(14)],
        'Kriteria Validitas': ["" for _ in range(14)]
    }
    df_media = pd.DataFrame(media_data)

    aspek_sistem = ['Fungsionalitas Sistem']*6 + ['Keandalan dan Validasi Sistem']*6 + ['Keamanan Akses dan Data yang Teramati']*5 + ['Kinerja Dasar Sistem']*5
    ind_sistem = [
        'Fungsi autentikasi (login/verifikasi)', 'Fungsi autentikasi (hak akses)', 'Fungsi proses layanan (alur pengajuan)', 
        'Fungsi proses layanan (form/berkas)', 'Fungsi proses layanan (revisi/tindak lanjut)', 'Kesesuaian fungsi inti',
        'Keandalan proses (status riwayat)', 'Keandalan proses (alur)', 'Validasi input (form)', 'Validasi input (unggah)', 
        'Penanganan kesalahan', 'Keandalan umum', 'Keamanan akses (halaman)', 'Keamanan akses (kewenangan)', 'Keamanan sesi', 
        'Keamanan data teramati', 'Keamanan unggah dan berkas', 'Kecepatan respons (halaman utama)', 'Kecepatan respons (transisi)', 
        'Stabilitas akses', 'Efisiensi penggunaan', 'Kualitas interaksi sistem'
    ]
    sistem_pak_ghea = [5,5,5,5,4,5, 5,4,5,4,4,5, 5,5,5,5,5, 5,5,4,5,5]
    sistem_pak_radinal = [4,5,5,4,4,5, 4,5,4,4,5,4, 3,3,4,4,4, 4,5,4,5,4]
    sistem_rahmad = [4,4,4,4,4,4, 4,4,5,5,4,5, 4,4,4,4,4, 4,4,4,4,4]
    
    sistem_data = {
        'Aspek Penilaian': aspek_sistem,
        'Indikator Penilaian': ind_sistem,
        'Ghea Chandra S': sistem_pak_ghea,
        'Radinal Fadli': sistem_pak_radinal,
        'Rahmad Wahyudi': sistem_rahmad,
        'Total Skor per Item': ["" for _ in range(22)],
        'Rata-rata Skor': ["" for _ in range(22)],
        'Skor Maks': ["" for _ in range(22)],
        'Persentase': ["" for _ in range(22)],
        'Kriteria Validitas': ["" for _ in range(22)]
    }
    df_sistem = pd.DataFrame(sistem_data)
    
    rekap_data = {
        'Aspek Penilaian': ['Ahli Materi', 'Ahli Media', 'Ahli Sistem', 'RATA-RATA KESELURUHAN'],
        'Total Skor': ["", "", "", ""],
        'Rata-rata Skor': ["", "", "", ""],
        'Skor Maksimal': ["", "", "", ""],
        'Persentase': ["", "", "", ""],
        'Kriteria Validitas': ["", "", "", ""]
    }
    df_rekap = pd.DataFrame(rekap_data)

    df_keputusan_gabungan = pd.DataFrame({
        'Nama Validator': ['Margaretha Karolina S', 'Eko Indra Pangestu', 'Putut Aji Nalendro', 
                           "",
                           'Rafiqa Rizalita', 'Daniel Rinaldi', 'Dwi Wahyudi',
                           "",
                           'Ghea Chandra S', 'Radinal Fadli', 'Rahmad Wahyudi',
                           ""],
        'Aspek': ['Materi', 'Materi', 'Materi', "", 'Media', 'Media', 'Media', "", 'Sistem', 'Sistem', 'Sistem', ""],
        'Keputusan Kelayakan': ['Layak digunakan dengan revisi', 'Layak digunakan tanpa revisi', 'Layak digunakan tanpa revisi',
                                "",
                                'Layak digunakan tanpa revisi', 'Layak digunakan tanpa revisi', 'Layak digunakan dengan revisi',
                                "",
                                'Layak digunakan tanpa revisi', 'Layak digunakan dengan revisi', 'Layak digunakan tanpa revisi',
                                ""],
        'Komentar/Saran': ['Tambah penjelasan, cek kalimat, pastikan screenshot jelas', '-', 'Beri pembatas header/content/footer & gambar',
                           "",
                           'Bagus, bisa digunakan untuk penelitian', 'Warna sebaiknya jangan gradasi', 'Beri icon pada setiap tile layanan',
                           "",
                           'Perbaikan saran', 'Login SSO, perhatikan CSP (unsafe-eval/inline), atur CSP default', 'Tampilkan tombol kembali ke atas',
                           ""]
    })

    filename = 'Rekap_Uji_Validitas_Ahli.xlsx'
    
    # Try writing to the main filename, fallback if locked
    try:
        with pd.ExcelWriter(filename) as writer:
            df_rekap.to_excel(writer, sheet_name='Rekap Keseluruhan', index=False)
            df_keputusan_gabungan.to_excel(writer, sheet_name='Rekap Keseluruhan', index=False, startrow=8)
            
            df_materi.to_excel(writer, sheet_name='Validitas Materi', index=False)
            df_media.to_excel(writer, sheet_name='Validitas Media', index=False)
            df_sistem.to_excel(writer, sheet_name='Validitas Sistem', index=False)
        apply_formatting_and_formulas(filename)
        print(f"Excel file '{filename}' with formulas and styling has been generated successfully.")
    except PermissionError:
        fallback_filename = 'Rekap_Uji_Validitas_Ahli_FINAL.xlsx'
        print(f"WARNING: '{filename}' is locked. Writing to '{fallback_filename}' instead.")
        with pd.ExcelWriter(fallback_filename) as writer:
            df_rekap.to_excel(writer, sheet_name='Rekap Keseluruhan', index=False)
            df_keputusan_gabungan.to_excel(writer, sheet_name='Rekap Keseluruhan', index=False, startrow=8)
            
            df_materi.to_excel(writer, sheet_name='Validitas Materi', index=False)
            df_media.to_excel(writer, sheet_name='Validitas Media', index=False)
            df_sistem.to_excel(writer, sheet_name='Validitas Sistem', index=False)
        apply_formatting_and_formulas(fallback_filename)
        print(f"Excel file '{fallback_filename}' has been generated successfully.")

if __name__ == '__main__':
    main()
